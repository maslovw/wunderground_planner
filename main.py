from openpyxl import Workbook
from openpyxl import load_workbook

def readFormCsv(file_name):
    database = {}
    with open(file_name, 'r', newline='') as csvfile:
        for line in csvfile:
            city_info = line.rstrip().split(';')
            city_name = city_info[0].encode("utf-8").strip()
            country_name = city_info[1].encode("utf-8").strip()
            key = city_name + country_name
            database[key] = city_info
    return database

def filldb(city_info, ws_out, i_out, key):
    ws_out["C" + str(i_out)] = city_info[key][2]
    ws_out["D" + str(i_out)] = city_info[key][3]
    ws_out["F" + str(i_out)] = city_info[key][4]
    ws_out["G" + str(i_out)] = city_info[key][5]
    ws_out["H" + str(i_out)] = city_info[key][6]
    ws_out["I" + str(i_out)] = city_info[key][7]
    ws_out["J" + str(i_out)] = city_info[key][8]
    ws_out["K" + str(i_out)] = city_info[key][9]
    ws_out["L" + str(i_out)] = city_info[key][10]
    ws_out["M" + str(i_out)] = city_info[key][11]
    ws_out["N" + str(i_out)] = city_info[key][12]
    ws_out["O" + str(i_out)] = city_info[key][13]
    ws_out["P" + str(i_out)] = city_info[key][14]
    ws_out["Q" + str(i_out)] = city_info[key][15]
    ws_out["R" + str(i_out)] = city_info[key][16]
    ws_out["S" + str(i_out)] = city_info[key][17]
    ws_out["T" + str(i_out)] = city_info[key][18]

wb = load_workbook("cities.xlsx")
ws = wb.active

wb_out = Workbook()
ws_out = wb_out.active
i_out = 1
i = 4
city_info = readFormCsv('db.csv')
if 1:
    while ws["A" + str(i)].value != None:
        print(ws['A' + str(i)].value, ws['B' + str(i)].value)
        ws_out["A" + str(i_out)] = ws['A' + str(i)].value.encode("utf-8").strip()
        ws_out["B" + str(i_out)] = ws['B' + str(i)].value
        try:
            city = ws['A' + str(i)].value.encode("utf-8").strip()
            country = ws['B' + str(i)].value.encode("utf-8").strip()
            key = city + country
            filldb(city_info, ws_out, i_out, key)
            city_info.pop(key)
        except:
            print ("SomeError")
        i_out += 1
        i += 1
print ("---------------------")
for item in city_info.keys():
    i_out += 1
    ws_out["A" + str(i_out)] = city_info[item][0]
    ws_out["B" + str(i_out)] = city_info[item][1]
    filldb(city_info, ws_out, i_out, item)
wb_out.save('db.xlsx')
# getYearWeather("Italy", "Cagliari")

