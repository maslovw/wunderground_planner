import csv
from openpyxl import Workbook
from openpyxl import load_workbook
from getWeatherFree import Wunderground
import json
import weather_error
import wbase

def writeCsv(udata, file_name ='db.csv'):
    with open(file_name, 'a', newline='') as csvfile:
        spamwriter = csv.writer(csvfile, delimiter=';')
        spamwriter.writerow(udata)

def readCsv(file_name='db.csv', off=0):
    with open(file_name, 'r', newline='') as csvfile:
        csvfile.seek(off)
        st = csvfile.readline()
        print(st.split(';'))
        i = csvfile.tell()
    return i

def write_data_csv(city_info, city, country):
    dec_str = 'ascii'
    udata = [
        city.decode(dec_str, 'ignore'), country.decode(dec_str, 'ignore'),
        city_info['city_name'],
        city_info['country_name'],
        city_info['lat'],
        city_info['lon'],
        city_info['temp1'],
        city_info['temp2'],
        city_info['temp3'],
        city_info['temp4'],
        city_info['temp5'],
        city_info['temp6'],
        city_info['temp8'],
        city_info['temp7'],
        city_info['temp9'],
        city_info['temp10'],
        city_info['temp11'],
        city_info['temp12'],
        city_info['city_url']
    ]
    writeCsv(udata)

def get_and_save_weather_wbase(country_name, city_name):
    print("wbase:", city_name, country_name)
    w = wbase.weatherbase()
    dec_str = 'ascii'
    city_info = w.get_weather(city_name.decode(dec_str, 'ignore'), country_name.decode(dec_str, 'ignore'))
    write_data_csv(city_info, city_name, country_name)

def get_and_save_weather(weather, country_name, city_name):
    dec_str = 'ascii'

    print("wunderground:", city_name, country_name)
    try:
        city_info = weather.get_year_weather(country_name, city_name)
        try:
            write_data_csv(city_info, city_name, country_name)
        except:
            print(" >> can't write to the file")
            return False
    except weather_error.KeyError as e:
        print("Key Error: ", e.value)
        raise e
    except NameError as e:
        print(" >> ", e)
        udata = [
            city_name.decode(dec_str, 'ignore'),
            country_name.decode(dec_str, 'ignore'),
            'location error',
            weather.url
        ]
        writeCsv(udata, 'unknown.csv')
        return False
    except:
        udata = [
            city_name.decode(dec_str, 'ignore'),
            country_name.decode(dec_str, 'ignore'),
            'unknown',
            weather.url
        ]
        writeCsv(udata, 'unknown.csv')
        print(" >> can't get", city_name, country_name)
        return False
    return True


def get_line_cfg():
    try:
        with open('app.json', 'r') as f:
            config = json.load(f)
    except:
        config = {'line': 1}
    try:
        return config['line']
    except:
        return 1

def save_cgf(line):
    config = {'line': line}
    with open('app.json', 'w') as f:
        json.dump(config, f)

def process_excel(file_name="160720_RIKA_CONTENT.xlsx"):
    wb = load_workbook(file_name)
    ws = wb.get_sheet_by_name("CITIES")

    i = get_line_cfg()
    weather = Wunderground()
    if 1:
        while ws["A" + str(i)].value != None:
            if ws["B" + str(i)].value == None:
                i += 1
                continue
            city_name = ws['A' + str(i)].value.encode("utf-8").strip()
            country_name = ws['B' + str(i)].value.encode("utf-8").strip()
            print(i, "req: ", weather.req_cnt)
            if not get_and_save_weather(weather, country_name, city_name):
                break
            i += 1
            save_cgf(i)

def process_csv(file_name="input.csv"):
    line_num = get_line_cfg()
    i = 0
    weather = Wunderground()
    with open(file_name, 'r', newline='') as csvfile:
        for line in csvfile:
            i += 1
            if i < line_num:
                continue
            city_info = line.rstrip().split(';')
            city_name = city_info[0].encode("utf-8").strip()
            country_name = city_info[1].encode("utf-8").strip()
            print(country_name, city_name)
            if country_name == "":
                continue
            print(i, "req: ", weather.req_cnt)

            try:
                if not get_and_save_weather(weather, country_name, city_name):
                    raise
            except:
                try:
                    get_and_save_weather_wbase(country_name, city_name)
                except:
                    print("no data")
            save_cgf(i)

process_csv("source.csv")
#wb_out.save('db.xlsx')